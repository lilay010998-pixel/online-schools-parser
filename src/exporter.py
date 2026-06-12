# Export module for saving data to Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Export school data to Excel format"""
    
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Schools"
    
    def _setup_styles(self):
        """Setup Excel styles"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        return {
            'header_fill': header_fill,
            'header_font': header_font,
            'header_alignment': header_alignment,
            'border': border
        }
    
    def export(self, schools_data: List[Dict]):
        """Export schools data to Excel"""
        try:
            styles = self._setup_styles()
            
            # Setup columns
            columns = [
                'Название компании',
                'Описание',
                'Сайт',
                'Email',
                'Телефон',
                'ВКонтакте',
                'Instagram',
                'Telegram',
                'YouTube',
                'Курсы'
            ]
            
            # Write headers
            for col_idx, col_name in enumerate(columns, 1):
                cell = self.ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
                cell.alignment = styles['header_alignment']
            
            # Set column widths
            column_widths = [25, 30, 25, 20, 15, 20, 20, 20, 20, 30]
            for col_idx, width in enumerate(column_widths, 1):
                self.ws.column_dimensions[chr(64 + col_idx)].width = width
            
            # Write data
            for row_idx, school in enumerate(schools_data, 2):
                cells = [
                    self.ws.cell(row=row_idx, column=1),
                    self.ws.cell(row=row_idx, column=2),
                    self.ws.cell(row=row_idx, column=3),
                    self.ws.cell(row=row_idx, column=4),
                    self.ws.cell(row=row_idx, column=5),
                    self.ws.cell(row=row_idx, column=6),
                    self.ws.cell(row=row_idx, column=7),
                    self.ws.cell(row=row_idx, column=8),
                    self.ws.cell(row=row_idx, column=9),
                    self.ws.cell(row=row_idx, column=10),
                ]
                
                values = [
                    school.get('name', ''),
                    school.get('description', ''),
                    school.get('url', ''),
                    school.get('contacts', {}).get('email', ''),
                    school.get('contacts', {}).get('phone', ''),
                    school.get('social_media', {}).get('vk', ''),
                    school.get('social_media', {}).get('instagram', ''),
                    school.get('social_media', {}).get('telegram', ''),
                    school.get('social_media', {}).get('youtube', ''),
                    ', '.join(school.get('courses', [])),
                ]
                
                for cell, value in zip(cells, values):
                    cell.value = value
                    cell.border = styles['border']
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Freeze first row
            self.ws.freeze_panes = "A2"
            
            # Save file
            self.wb.save(self.filename)
            logger.info(f"✅ Excel файл сохранён: {self.filename}")
            logger.info(f"📊 Всего записей: {len(schools_data)}")
            
        except Exception as e:
            logger.error(f"❌ Error exporting to Excel: {str(e)}")
            raise
